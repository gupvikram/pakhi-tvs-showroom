"""
Pakhi TVS Alwar Showroom – FastAPI Backend
Real-time showroom workflow: quotation → approval → payment → export
Port: 8001
"""

import asyncio
import json
import os
from datetime import datetime, date, timezone
from pathlib import Path
from typing import Optional, List, AsyncGenerator

import aiofiles
import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from supabase import create_client, Client

load_dotenv(override=True)

# ─── Directories ───────────────────────────────────────────────────────────────
EXPORTS_DIR = Path(__file__).parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)
FRONTEND_DIR = Path(__file__).parent / "frontend"

# ─── FastAPI App ───────────────────────────────────────────────────────────────
app = FastAPI(title="Pakhi TVS Alwar Showroom", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIR / "assets")), name="assets")
app.mount("/exports", StaticFiles(directory=str(EXPORTS_DIR)), name="exports_static")

# ─── Constants ────────────────────────────────────────────────────────────────
OWNER_PIN = os.getenv("OWNER_PIN", "1234")
APP_BASE_URL = os.getenv("APP_BASE_URL", "http://localhost:8001")
HSN_TWO_WHEELER = "87112019"

# ─── Supabase client ────────────────────────────────────────────────────────
# Fresh client per request — avoids stale httpx connection pool on macOS (errno 35)
def get_db() -> Client:
    url = os.getenv("SUPABASE_URL", "")
    key = os.getenv("SUPABASE_KEY", "")
    if not url or not key:
        raise HTTPException(500, detail="SUPABASE_URL and SUPABASE_KEY must be set in .env")
    return create_client(url, key)

# ─── SSE broadcast system ─────────────────────────────────────────────────────
_sse_queues: list[asyncio.Queue] = []
_main_loop: Optional[asyncio.AbstractEventLoop] = None

@app.on_event("startup")
async def startup():
    global _main_loop, _db
    _main_loop = asyncio.get_event_loop()
    _db = None  # Always create a fresh Supabase client on startup/reload

def sync_broadcast(event_type: str, data: dict):
    """Thread-safe broadcast from sync handlers to all SSE clients."""
    if not _main_loop:
        return
    payload = f"event: {event_type}\ndata: {json.dumps(data)}\n\n"
    for q in list(_sse_queues):
        try:
            _main_loop.call_soon_threadsafe(q.put_nowait, payload)
        except Exception:
            pass

# ─── Road tax (Rajasthan) ─────────────────────────────────────────────────────
def calculate_road_tax(ex_showroom_paise: int, engine_cc: int, fuel_type: str):
    """Returns (road_tax_paise: int, rate: float). All EVs are exempt."""
    if fuel_type == "electric":
        return 0, 0.0
    price_rupees = ex_showroom_paise / 100
    if engine_cc <= 125:
        rate = 0.04
    elif engine_cc <= 200:
        rate = 0.06
    elif engine_cc <= 500:
        rate = 0.08
    else:
        rate = 0.10
    road_tax = int(price_rupees * rate * 100)
    return road_tax, rate

# ─── Quote / Invoice number generation ────────────────────────────────────────
def _next_sequence_number(db: Client, year: int, prefix: str, field: str) -> str:
    """Count existing records for the year and return the next formatted number."""
    start = f"{year}-01-01T00:00:00+00:00"
    end   = f"{year+1}-01-01T00:00:00+00:00"
    result = db.table("quotations").select("id").gte("created_at", start).lt("created_at", end).execute()
    count = len(result.data) + 1
    return f"{prefix}-{year}-{count:04d}"

def _next_invoice_number(db: Client, year: int) -> str:
    start = f"{year}-01-01T00:00:00+00:00"
    end   = f"{year+1}-01-01T00:00:00+00:00"
    result = (
        db.table("quotations")
        .select("id")
        .not_.is_("invoice_number", "null")
        .gte("created_at", start)
        .lt("created_at", end)
        .execute()
    )
    count = len(result.data) + 1
    return f"INV-{year}-{count:04d}"

# ─── Telegram notifications ───────────────────────────────────────────────────
async def _send_telegram(message: str) -> None:
    token = os.getenv("TELEGRAM_BOT_TOKEN", "")
    chat_id = os.getenv("TELEGRAM_OWNER_CHAT_ID", "")
    if not token or not chat_id:
        return
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": chat_id, "text": message, "parse_mode": "HTML"},
            )
    except Exception:
        pass

def _schedule_telegram(bg: BackgroundTasks, message: str) -> None:
    bg.add_task(_send_telegram, message)

# ─── HTML page routes ──────────────────────────────────────────────────────────
async def _html(name: str) -> HTMLResponse:
    path = FRONTEND_DIR / name
    if not path.exists():
        raise HTTPException(404, detail=f"{name} not found")
    async with aiofiles.open(path, encoding="utf-8") as f:
        return HTMLResponse(await f.read())

@app.get("/",        include_in_schema=False)
async def root():        return await _html("sales.html")

@app.get("/sales",   include_in_schema=False)
async def sales():   return await _html("sales.html")

@app.get("/approve", include_in_schema=False)
async def approve(): return await _html("approve.html")

@app.get("/cashier", include_in_schema=False)
async def cashier(): return await _html("cashier.html")

@app.get("/accounts",include_in_schema=False)
async def accounts():return await _html("accounts.html")

@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    # Tiny TVS-red SVG favicon — silences browser 404 warnings
    svg = b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32"><rect width="32" height="32" rx="6" fill="#C41E3A"/><text x="50%" y="54%" dominant-baseline="middle" text-anchor="middle" font-size="18" font-family="Arial" font-weight="bold" fill="#fff">T</text></svg>'
    from fastapi.responses import Response
    return Response(content=svg, media_type="image/svg+xml")

# ─── Catalog ──────────────────────────────────────────────────────────────────
@app.get("/catalog")
def get_catalog():
    db = get_db()
    vehicles_res = db.table("vehicles").select(
        "id, brand, model, category, variants(id, name, ex_showroom_price, engine_cc, fuel_type)"
    ).eq("active", True).order("category").order("model").execute()

    catalog: dict = {}
    for v in (vehicles_res.data or []):
        cat = v["category"]
        if cat not in catalog:
            catalog[cat] = []
        vehicle_entry = {
            "id": v["id"],
            "brand": v["brand"],
            "model": v["model"],
            "variants": [],
        }
        for variant in (v.get("variants") or []):
            if not variant.get("active", True):
                continue
            rt_paise, rt_rate = calculate_road_tax(
                variant["ex_showroom_price"],
                variant.get("engine_cc", 0) or 0,
                variant.get("fuel_type", "petrol"),
            )
            vehicle_entry["variants"].append({
                "id": variant["id"],
                "name": variant["name"],
                "ex_showroom_price": variant["ex_showroom_price"],
                "engine_cc": variant.get("engine_cc", 0),
                "fuel_type": variant.get("fuel_type", "petrol"),
                "road_tax_amount": rt_paise,
                "road_tax_rate": rt_rate,
            })
        if vehicle_entry["variants"]:
            catalog[cat].append(vehicle_entry)

    return {"catalog": catalog}

# ─── Reps ────────────────────────────────────────────────────────────────────
@app.get("/reps")
def get_reps():
    db = get_db()
    result = db.table("reps").select("*").eq("active", True).order("id").execute()
    return {"reps": result.data or []}

# ─── Accessories ──────────────────────────────────────────────────────────────
@app.get("/accessories")
def get_accessories():
    db = get_db()
    result = db.table("accessories").select("*").eq("active", True).order("category").order("name").execute()
    grouped: dict = {}
    for acc in (result.data or []):
        cat = acc["category"]
        if cat not in grouped:
            grouped[cat] = []
        grouped[cat].append(acc)
    return {"accessories": grouped}

# ─── Finance Companies ────────────────────────────────────────────────────────
@app.get("/finance-companies")
def get_finance_companies():
    db = get_db()
    result = db.table("finance_companies").select("*").eq("active", True).order("name").execute()
    return {"finance_companies": result.data or []}

# ─── Pydantic models ─────────────────────────────────────────────────────────
class AccessoryItem(BaseModel):
    accessory_id: int
    name: str
    price_at_time: int  # paise

class CreateQuotationRequest(BaseModel):
    rep_id: int
    customer_name: str
    customer_phone: str
    customer_address: str
    variant_id: int
    registration_fee: int = 40000
    hsrp_charges: int = 45000
    smart_card_rc: int = 20000
    fasstag: int = 50000
    insurance_premium: int = 0
    insurance_company: Optional[str] = None
    insurance_years: int = 1
    finance_company_id: Optional[int] = None
    hypothecation_charges: int = 0
    extended_warranty_price: int = 0
    extended_warranty_years: int = 0
    accessories: List[AccessoryItem] = []

class ApproveQuotationRequest(BaseModel):
    pin: str
    discount_amount: int = 0
    discount_reason: Optional[str] = None

class RejectQuotationRequest(BaseModel):
    pin: str
    rejection_reason: str

class PayQuotationRequest(BaseModel):
    payment_mode: str  # Cash / UPI / Finance
    upi_reference: Optional[str] = None
    finance_company_id: Optional[int] = None
    finance_reference: Optional[str] = None

class ExportRequest(BaseModel):
    quotation_ids: List[int]

# ─── Quotation helpers ────────────────────────────────────────────────────────
def _enrich_quotation(q: dict, db: Client) -> dict:
    """Attach rep, variant, vehicle, finance company, and accessories to a quotation row."""
    if q.get("rep_id"):
        rep = db.table("reps").select("id, name, emoji").eq("id", q["rep_id"]).single().execute()
        q["rep"] = rep.data
    if q.get("variant_id"):
        var = db.table("variants").select("id, name, ex_showroom_price, engine_cc, fuel_type, vehicles(id, brand, model, category)").eq("id", q["variant_id"]).single().execute()
        q["variant"] = var.data
    if q.get("finance_company_id"):
        fc = db.table("finance_companies").select("id, name").eq("id", q["finance_company_id"]).single().execute()
        q["finance_company"] = fc.data
    acc_res = db.table("quotation_accessories").select("*").eq("quotation_id", q["id"]).execute()
    q["accessories_list"] = acc_res.data or []
    return q

def _build_event_payload(q: dict) -> dict:
    variant = q.get("variant") or {}
    vehicle = (variant.get("vehicles") or {}) if isinstance(variant, dict) else {}
    rep = q.get("rep") or {}
    brand = vehicle.get("brand", "")
    model = vehicle.get("model", "")
    vname = variant.get("name", "")
    return {
        "quotation_id": q["id"],
        "status": q["status"],
        "customer_name": q["customer_name"],
        "vehicle_name": f"{brand} {model} {vname}".strip(),
        "final_amount": q["final_amount"],
        "gross_amount": q["gross_amount"],
        "quote_number": q.get("quote_number"),
        "invoice_number": q.get("invoice_number"),
        "rep_name": rep.get("name", ""),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }

# ─── POST /quotations ─────────────────────────────────────────────────────────
@app.post("/quotations", status_code=201)
def create_quotation(req: CreateQuotationRequest, bg: BackgroundTasks):
    db = get_db()

    # Fetch variant to get ex_showroom_price and engine specs
    var_res = db.table("variants").select("id, name, ex_showroom_price, engine_cc, fuel_type, vehicles(id, brand, model)").eq("id", req.variant_id).single().execute()
    if not var_res.data:
        raise HTTPException(404, detail="Variant not found")
    variant = var_res.data
    vehicle = variant.get("vehicles") or {}

    ex_price = variant["ex_showroom_price"]
    rt_paise, rt_rate = calculate_road_tax(ex_price, variant.get("engine_cc", 0) or 0, variant.get("fuel_type", "petrol"))

    accessories_total = sum(a.price_at_time for a in req.accessories)
    gross_amount = (
        ex_price
        + rt_paise
        + req.registration_fee
        + req.hsrp_charges
        + req.smart_card_rc
        + req.fasstag
        + req.insurance_premium
        + req.hypothecation_charges
        + req.extended_warranty_price
        + accessories_total
    )

    year = datetime.now().year
    quote_number = _next_sequence_number(db, year, "Q", "quote_number")

    row = {
        "quote_number": quote_number,
        "rep_id": req.rep_id,
        "customer_name": req.customer_name,
        "customer_phone": req.customer_phone,
        "customer_address": req.customer_address,
        "variant_id": req.variant_id,
        "ex_showroom_price": ex_price,
        "road_tax_rate": rt_rate,
        "road_tax_amount": rt_paise,
        "registration_fee": req.registration_fee,
        "hsrp_charges": req.hsrp_charges,
        "smart_card_rc": req.smart_card_rc,
        "fasstag": req.fasstag,
        "insurance_premium": req.insurance_premium,
        "insurance_company": req.insurance_company,
        "insurance_years": req.insurance_years,
        "hypothecation_charges": req.hypothecation_charges,
        "finance_company_id": req.finance_company_id,
        "extended_warranty_price": req.extended_warranty_price,
        "extended_warranty_years": req.extended_warranty_years,
        "accessories_total": accessories_total,
        "gross_amount": gross_amount,
        "final_amount": gross_amount,
        "status": "pending_approval",
    }

    ins = db.table("quotations").insert(row).execute()
    if not ins.data:
        raise HTTPException(500, detail="Failed to create quotation")

    quotation = ins.data[0]
    qid = quotation["id"]

    # Insert accessories
    if req.accessories:
        acc_rows = [
            {"quotation_id": qid, "accessory_id": a.accessory_id, "name": a.name, "price_at_time": a.price_at_time}
            for a in req.accessories
        ]
        db.table("quotation_accessories").insert(acc_rows).execute()

    # Fetch rep info for Telegram
    rep_name = ""
    rep_res = db.table("reps").select("name").eq("id", req.rep_id).single().execute()
    if rep_res.data:
        rep_name = rep_res.data["name"]

    brand = vehicle.get("brand", "")
    model_name = vehicle.get("model", "")
    variant_name = variant.get("name", "")
    gross_rupees = gross_amount // 100

    msg = (
        f"🔔 <b>New Quote Pending Approval</b>\n\n"
        f"👤 Customer: {req.customer_name}\n"
        f"📱 Phone: {req.customer_phone}\n"
        f"🏍️ Vehicle: {brand} {model_name} - {variant_name}\n"
        f"💰 Total: ₹{gross_rupees:,}\n"
        f"👨‍💼 Sales Rep: {rep_name}\n"
        f"🔖 Quote #: {quote_number}\n\n"
        f"Tap to approve: {APP_BASE_URL}/approve"
    )
    _schedule_telegram(bg, msg)

    sync_broadcast("quotation_submitted", {
        "quotation_id": qid,
        "status": "pending_approval",
        "customer_name": req.customer_name,
        "vehicle_name": f"{brand} {model_name} {variant_name}".strip(),
        "gross_amount": gross_amount,
        "final_amount": gross_amount,
        "quote_number": quote_number,
        "invoice_number": None,
        "rep_name": rep_name,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })

    return {"quotation": quotation, "quote_number": quote_number}

# ─── GET /quotations ───────────────────────────────────────────────────────────
@app.get("/quotations")
def list_quotations(
    status: Optional[str] = Query(None),
    date_filter: Optional[str] = Query(None, alias="date"),
):
    db = get_db()
    q = db.table("quotations").select(
        "*, reps(id, name, emoji), variants(id, name, ex_showroom_price, fuel_type, vehicles(brand, model, category)), finance_companies(id, name)"
    ).order("created_at", desc=True)

    if status:
        q = q.eq("status", status)

    now = datetime.now(timezone.utc)
    if date_filter == "today":
        day_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        q = q.gte("created_at", day_start)
    elif date_filter == "week":
        week_start = (now.replace(hour=0, minute=0, second=0, microsecond=0)).isoformat()
        # approx 7 days
        from datetime import timedelta
        week_ago  = (now - timedelta(days=7)).isoformat()
        q = q.gte("created_at", week_ago)
    elif date_filter == "month":
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        q = q.gte("created_at", month_start)

    result = q.execute()
    return {"quotations": result.data or []}

# ─── GET /quotations/{id} ──────────────────────────────────────────────────────
@app.get("/quotations/{quotation_id}")
def get_quotation(quotation_id: int):
    db = get_db()
    res = db.table("quotations").select("*").eq("id", quotation_id).single().execute()
    if not res.data:
        raise HTTPException(404, detail="Quotation not found")
    q = _enrich_quotation(res.data, db)
    return {"quotation": q}

# ─── POST /quotations/{id}/approve ────────────────────────────────────────────
@app.post("/quotations/{quotation_id}/approve")
def approve_quotation(quotation_id: int, req: ApproveQuotationRequest, bg: BackgroundTasks):
    if req.pin != OWNER_PIN:
        raise HTTPException(401, detail="Incorrect PIN")

    db = get_db()
    res = db.table("quotations").select("*").eq("id", quotation_id).single().execute()
    if not res.data:
        raise HTTPException(404, detail="Quotation not found")
    q = res.data

    if q["status"] not in ("pending_approval",):
        raise HTTPException(400, detail=f"Quotation is already {q['status']}")

    discount = req.discount_amount or 0
    final_amount = q["gross_amount"] - discount
    if final_amount < 0:
        raise HTTPException(400, detail="Discount cannot exceed the gross amount")

    year = datetime.now().year
    invoice_number = _next_invoice_number(db, year)

    update_data = {
        "status": "approved",
        "invoice_number": invoice_number,
        "discount_amount": discount,
        "discount_reason": req.discount_reason,
        "final_amount": final_amount,
        "approved_at": datetime.now(timezone.utc).isoformat(),
    }
    updated = db.table("quotations").update(update_data).eq("id", quotation_id).execute()
    q.update(update_data)

    enriched = _enrich_quotation(q, db)
    variant  = enriched.get("variant") or {}
    vehicle  = (variant.get("vehicles") or {}) if isinstance(variant, dict) else {}
    brand    = vehicle.get("brand", "")
    model_n  = vehicle.get("model", "")
    rep_name = (enriched.get("rep") or {}).get("name", "")

    final_rupees = final_amount // 100
    msg = (
        f"✅ <b>Quote Approved</b>\n\n"
        f"🧾 Invoice #: {invoice_number}\n"
        f"👤 Customer: {q['customer_name']}\n"
        f"🏍️ Vehicle: {brand} {model_n} - {variant.get('name','')}\n"
        f"💰 Final Amount: ₹{final_rupees:,}\n"
    )
    if discount > 0:
        msg += (
            f"🏷️ Discount Applied: ₹{discount//100:,}\n"
            f"📝 Reason: {req.discount_reason or 'N/A'}\n"
        )
    msg += "📢 Sales rep and cashier notified."
    _schedule_telegram(bg, msg)

    sync_broadcast("quotation_approved", _build_event_payload(enriched))
    return {"quotation": updated.data[0] if updated.data else q, "invoice_number": invoice_number}

# ─── POST /quotations/{id}/reject ─────────────────────────────────────────────
@app.post("/quotations/{quotation_id}/reject")
def reject_quotation(quotation_id: int, req: RejectQuotationRequest, bg: BackgroundTasks):
    if req.pin != OWNER_PIN:
        raise HTTPException(401, detail="Incorrect PIN")
    if not req.rejection_reason or not req.rejection_reason.strip():
        raise HTTPException(400, detail="rejection_reason is required")

    db = get_db()
    res = db.table("quotations").select("*").eq("id", quotation_id).single().execute()
    if not res.data:
        raise HTTPException(404, detail="Quotation not found")
    q = res.data

    updated = db.table("quotations").update({
        "status": "rejected",
        "rejection_reason": req.rejection_reason,
    }).eq("id", quotation_id).execute()
    q["status"] = "rejected"

    enriched = _enrich_quotation(q, db)
    variant  = enriched.get("variant") or {}
    vehicle  = (variant.get("vehicles") or {}) if isinstance(variant, dict) else {}
    brand    = vehicle.get("brand", "")
    model_n  = vehicle.get("model", "")

    msg = (
        f"❌ <b>Quote Rejected</b>\n\n"
        f"👤 Customer: {q['customer_name']}\n"
        f"🏍️ Vehicle: {brand} {model_n}\n"
        f"❓ Reason: {req.rejection_reason}\n"
        f"📢 Sales rep has been notified."
    )
    _schedule_telegram(bg, msg)
    sync_broadcast("quotation_rejected", _build_event_payload(enriched))
    return {"message": "Quotation rejected", "quotation_id": quotation_id}

# ─── POST /quotations/{id}/pay ────────────────────────────────────────────────
@app.post("/quotations/{quotation_id}/pay")
def pay_quotation(quotation_id: int, req: PayQuotationRequest, bg: BackgroundTasks):
    if req.payment_mode == "UPI" and not req.upi_reference:
        raise HTTPException(400, detail="upi_reference is required for UPI payments")
    if req.payment_mode == "Finance" and not req.finance_reference:
        raise HTTPException(400, detail="finance_reference is required for Finance payments")

    db = get_db()
    res = db.table("quotations").select("*").eq("id", quotation_id).single().execute()
    if not res.data:
        raise HTTPException(404, detail="Quotation not found")
    q = res.data

    if q["status"] != "approved":
        raise HTTPException(400, detail=f"Quotation status is {q['status']}, expected approved")

    now_iso = datetime.now(timezone.utc).isoformat()
    update_data = {
        "status": "paid",
        "payment_mode": req.payment_mode,
        "upi_reference": req.upi_reference,
        "finance_reference": req.finance_reference,
        "finance_company_id": req.finance_company_id or q.get("finance_company_id"),
        "paid_at": now_iso,
    }
    db.table("quotations").update(update_data).eq("id", quotation_id).execute()

    # Create payment record
    db.table("payments").insert({
        "quotation_id": quotation_id,
        "amount": q["final_amount"],
        "payment_mode": req.payment_mode,
        "upi_reference": req.upi_reference,
        "finance_company_id": req.finance_company_id or q.get("finance_company_id"),
        "finance_reference": req.finance_reference,
    }).execute()

    q.update(update_data)
    enriched = _enrich_quotation(q, db)
    sync_broadcast("quotation_paid", _build_event_payload(enriched))
    return {"message": "Payment recorded", "quotation_id": quotation_id}

# ─── POST /quotations/export ──────────────────────────────────────────────────
@app.post("/quotations/export")
def export_quotations(req: ExportRequest):
    db = get_db()

    rows_data = []
    for qid in req.quotation_ids:
        res = db.table("quotations").select("*").eq("id", qid).eq("status", "paid").single().execute()
        if not res.data:
            continue
        q = _enrich_quotation(res.data, db)
        rows_data.append(q)

    if not rows_data:
        raise HTTPException(400, detail="No paid quotations found for the provided IDs")

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vouchers"

    COLUMNS = [
        "Vch_Series", "Date", "Invoice_No", "Party_Name", "Party_Phone",
        "Party_Address", "Vehicle_Model", "Variant_Name", "HSN_Code", "Qty",
        "Ex_Showroom_Price", "Road_Tax", "Registration_Charges", "Insurance_Premium",
        "Insurance_Company", "Accessories_Total", "Gross_Amount", "Discount_Amount",
        "Final_Amount", "Payment_Mode", "Finance_Company", "Finance_Reference",
        "UPI_Reference", "Sales_Rep", "Narration",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, q in enumerate(rows_data, 2):
        variant = q.get("variant") or {}
        vehicle = (variant.get("vehicles") or {}) if isinstance(variant, dict) else {}
        rep     = q.get("rep") or {}
        fc      = q.get("finance_company") or {}

        paid_at  = q.get("paid_at") or q.get("approved_at") or q.get("created_at") or ""
        try:
            dt = datetime.fromisoformat(paid_at.replace("Z", "+00:00"))
            date_str = dt.strftime("%d-%m-%Y")
        except Exception:
            date_str = ""

        brand      = vehicle.get("brand", "")
        model_name = vehicle.get("model", "")
        variant_name = variant.get("name", "") if isinstance(variant, dict) else ""
        rep_name   = rep.get("name", "") if isinstance(rep, dict) else ""
        fc_name    = fc.get("name", "") if isinstance(fc, dict) else ""

        def p2r(paise):
            """Convert paise to rupees (2 decimal float)."""
            return round((paise or 0) / 100, 2)

        narration = (
            f"Vehicle Sale: {brand} {model_name} {variant_name}, "
            f"Invoice {q.get('invoice_number','')}, HSN {HSN_TWO_WHEELER}"
        )

        values = [
            "Sales",
            date_str,
            q.get("invoice_number", ""),
            q.get("customer_name", ""),
            q.get("customer_phone", ""),
            q.get("customer_address", ""),
            f"{brand} {model_name}".strip(),
            variant_name,
            HSN_TWO_WHEELER,
            1,
            p2r(q.get("ex_showroom_price")),
            p2r(q.get("road_tax_amount")),
            p2r((q.get("registration_fee",0) or 0) + (q.get("hsrp_charges",0) or 0) + (q.get("smart_card_rc",0) or 0) + (q.get("fasstag",0) or 0)),
            p2r(q.get("insurance_premium")),
            q.get("insurance_company", ""),
            p2r(q.get("accessories_total")),
            p2r(q.get("gross_amount")),
            p2r(q.get("discount_amount")),
            p2r(q.get("final_amount")),
            q.get("payment_mode", ""),
            fc_name,
            q.get("finance_reference", "") or "",
            q.get("upi_reference", "") or "",
            rep_name,
            narration,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin

    # Auto-size columns
    for col_idx, col_name in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row_idx in range(2, len(rows_data) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    today = date.today().strftime("%Y-%m-%d")
    filename = f"TVS_Alwar_Export_{today}.xlsx"
    output_path = EXPORTS_DIR / filename
    wb.save(str(output_path))

    # Mark as exported
    now_iso = datetime.now(timezone.utc).isoformat()
    for q in rows_data:
        db.table("quotations").update({"busy_exported_at": now_iso}).eq("id", q["id"]).execute()

    return FileResponse(
        path=str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ─── GET /events (SSE) ────────────────────────────────────────────────────────
@app.get("/events")
async def sse_endpoint():
    async def generator() -> AsyncGenerator[str, None]:
        q: asyncio.Queue = asyncio.Queue(maxsize=200)
        _sse_queues.append(q)
        try:
            while True:
                try:
                    event = await asyncio.wait_for(q.get(), timeout=15)
                    yield event
                except asyncio.TimeoutError:
                    yield ": ping\n\n"
        finally:
            try:
                _sse_queues.remove(q)
            except ValueError:
                pass

    return StreamingResponse(
        generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

# ─── Stats endpoint ───────────────────────────────────────────────────────────
@app.get("/stats")
def get_stats(date_filter: Optional[str] = Query("today", alias="date")):
    db = get_db()
    now = datetime.now(timezone.utc)
    from datetime import timedelta

    if date_filter == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    elif date_filter == "week":
        start = (now - timedelta(days=7)).isoformat()
    elif date_filter == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    elif date_filter == "all":
        start = "2000-01-01T00:00:00+00:00"
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()

    # Quotations in period
    all_q = db.table("quotations").select(
        "id, status, gross_amount, final_amount, discount_amount, rep_id, busy_exported_at, reps(name, emoji), created_at"
    ).gte("created_at", start).execute()
    rows = all_q.data or []

    paid_rows    = [r for r in rows if r["status"] in ("paid", "exported")]
    pending_exp  = [r for r in rows if r["status"] == "paid" and not r.get("busy_exported_at")]
    discounts    = sum((r.get("discount_amount") or 0) for r in paid_rows)

    # Pending approvals — always ALL TIME (not period-filtered, so nothing slips through)
    pa_res = db.table("quotations").select(
        "id, gross_amount, rep_id, reps(name, emoji)"
    ).eq("status", "pending_approval").execute()
    pending_rows_all = pa_res.data or []
    pending_approvals = len(pending_rows_all)

    # Approved but not yet paid — always ALL TIME
    appr_res = db.table("quotations").select(
        "id, final_amount, rep_id, reps(name, emoji)"
    ).eq("status", "approved").execute()
    approved_rows_all = appr_res.data or []

    # Reps with pending quotes (for "Awaiting Approval" section)
    rep_pending: dict = {}
    for r in pending_rows_all:
        rep  = r.get("reps") or {}
        rid  = r.get("rep_id")
        if rid not in rep_pending:
            rep_pending[rid] = {
                "rep_name":  rep.get("name", ""),
                "rep_emoji": rep.get("emoji", ""),
                "count":     0,
                "amount":    0,
            }
        rep_pending[rid]["count"]  += 1
        rep_pending[rid]["amount"] += r.get("gross_amount", 0)
    reps_with_pending = sorted(rep_pending.values(), key=lambda x: -x["amount"])

    # Leaderboard — pull ALL statuses for the period so we can show full pipeline per rep
    rep_stats: dict = {}
    for r in rows:
        rep = (r.get("reps") or {})
        rid = r.get("rep_id")
        if rid not in rep_stats:
            rep_stats[rid] = {
                "rep_name":    rep.get("name", ""),
                "rep_emoji":   rep.get("emoji", ""),
                "name":        rep.get("name", ""),   # legacy
                "emoji":       rep.get("emoji", ""),  # legacy
                "sales_count": 0,   # paid + exported — used for ranking
                "count":       0,   # legacy alias
                "pending_count":  0,
                "approved_count": 0,
                "paid_count":     0,
                "total_value": 0,
                "total":       0,   # legacy alias
            }
        st = r.get("status", "")
        if st == "pending_approval":
            rep_stats[rid]["pending_count"]  += 1
        elif st == "approved":
            rep_stats[rid]["approved_count"] += 1
        elif st in ("paid", "exported"):
            rep_stats[rid]["paid_count"]     += 1
            rep_stats[rid]["sales_count"]    += 1
            rep_stats[rid]["count"]          += 1
            rep_stats[rid]["total_value"]    += r.get("final_amount", 0)
            rep_stats[rid]["total"]          += r.get("final_amount", 0)

    leaderboard = sorted(rep_stats.values(), key=lambda x: (-x["sales_count"], -x["total_value"]))

    return {
        # ── New widget fields ──
        "total_sales":               len(paid_rows),
        "total_revenue":             sum(r.get("final_amount", 0) for r in paid_rows),
        "total_discounts":           discounts,
        "pending_approvals":         pending_approvals,
        "pending_approval_amount":   sum(r.get("gross_amount", 0) for r in pending_rows_all),
        "approved_amount":           sum(r.get("final_amount", 0) for r in approved_rows_all),
        "reps_with_pending":         reps_with_pending,
        "leaderboard":               leaderboard,
        # ── Legacy fields (accounts.html backward compat) ──
        "sales_count":          len(paid_rows),
        "sales_value":          sum(r.get("final_amount", 0) for r in paid_rows),
        "pending_export_count": len(pending_exp),
        "rep_leaderboard":      leaderboard,
    }

# ─── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8001, reload=True)
